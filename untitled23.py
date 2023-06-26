
import pytesseract
import PyPDF2
from PIL import Image
from pdf2image import convert_from_path
import re
import os
import pandas as pd
import openpyxl
pytesseract.pytesseract.tesseract_cmd = R'C:\Users\gabriel.rodrigues\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'


def verificar_correspondencia(elemento, numero):
    def buscar_texto(texto_extraido):
        termo = 'PROPOSTA DE COMPRA E VENDA'
        match = re.search(termo, texto_extraido)
        
        if match:
            return 1
        else:
            return 0
    
    def buscar_nome(nome_cliente, texto_extraido):
        
        match = re.search(nome_cliente, texto_extraido)
        
        if match:
            return 1
        else:
            return 0
    
    def buscar_caracter1(conjunto_palavras):
        termo = 'QUINTA'
        termo2 = 'extrajudicial'
        termo3 = 'honorarios'
        termo4 = 'advocaticios'
        
        for frase in conjunto_palavras:
            match = re.search(termo, frase)
            match2 = re.search(termo2, frase)
            match3 = re.search(termo3, frase)
            match4 = re.search(termo4, frase)
            if match:
                return 1
            elif match2:
                return 1
            elif match3:
                return 1
            elif match4:
                return 1
        return 0
        
    def buscar_caracter2(conjunto_palavras):
        termo = 'incorridos'
        termo2 = 'EXCLUSIVO'
        termo3 = '(cinco)'
        termo4 = 'Indivisibilidade'
        
        for frase in conjunto_palavras:
            match = re.search(termo, frase)
            match2 = re.search(termo2, frase)
            match3 = re.search(termo3, frase)
            match4 = re.search(termo4, frase)
            if match:
                return 1
            elif match2:
                return 1
            elif match3:
                return 1
            elif match4:
                return 1
        return 0
        
    
    i=0
    i3=0
    quantidade_paginas = 0

    caminho_3 = R'{}'.format(elemento)
    try:
        with open(caminho_3, "rb") as arquivo_pdf:
            # Faça algo com o arquivo
            pdfReader = PyPDF2.PdfReader(arquivo_pdf)
            quantidade_paginas = quantidade_paginas + int(len(pdfReader.pages))
    except FileNotFoundError:
        print('erro na leitura do pdf')
        
        return None
        # Lida com o erro de arquivo não encontrado
        
    termo = '.pdf'
    termo2 = 'image'
    indicador1 = 0
    indicador2 = 0
    indicador3 = 0
    indicador4 = 0
    indicador5 = 0
    verificar_texto = 0
    images = convert_from_path(caminho_3, first_page=0, last_page= quantidade_paginas )
    caminho_imagem = elemento.replace(termo, termo2)
    if quantidade_paginas < 10 and quantidade_paginas >= 1:
        return None
    
    if images:
        while i < quantidade_paginas:
            image = images[i]
            image.save(caminho_imagem, "JPEG")
            imagem = Image.open(caminho_imagem)
            texto_extraido = pytesseract.image_to_string(imagem)
            i = i + 1
            if indicador1 == 0:
                verificar_texto = buscar_texto(texto_extraido)
                if verificar_texto == 1:
                    
                    indicador1 = indicador1 + 1
            if indicador2 == 0 and i <11:
                      
                for i2 in range(0, len(nova_lista), 2):
                       
                    nome_cliente = nova_lista[i2]
                    if nome_cliente != 'nan':
                        verificar_nome = buscar_nome(nome_cliente, texto_extraido)
                        i3 = i2 +1
                        numero_cota = nova_lista[i3]
                        
                        if verificar_nome == 1 and numero_cota == numero:
                            indicador2 = indicador2 + 1
                            indicador3 = indicador3 +1
                            
                            break
            
        
            if i > 5 and indicador2 == 0 and indicador3 == 0:
                os.remove(caminho_imagem) 
                return None
            conjunto_palavras = set(texto_extraido.split(' '))
            if indicador4 == 0:
                
                verificar_caracter1 = buscar_caracter1(conjunto_palavras)
                
                if verificar_caracter1 == 1:
                    
                    indicador4 = indicador4 + 1
                    
            if indicador5 == 0:
                
                verificar_caracter2 = buscar_caracter2(conjunto_palavras)
                
                if verificar_caracter2 == 1:
                    
                    indicador5 = indicador5 + 1
            #print(verificar_texto, indicador1, indicador2, indicador3, indicador4, indicador5)        
            if indicador3 == 1 and indicador4 == 1 and indicador5 == 1 and indicador2 == 1:
                
                os.remove(caminho_imagem)                                      
                return nome_cliente +' + '+ numero_cota + ' + ' + str(quantidade_paginas)
                        
            else:
                
                continue
            
        else:
            i = i + 1
        
    os.remove(caminho_imagem)       
    return None
    
                    
# Lê o arquivo Excel com as duas colunas
df = pd.read_excel(R'C:\Users\gabriel.rodrigues\Desktop\PLANILHA_ARQUIVAR_TESTE.xlsx')
df2 = pd.read_excel(R'C:\Users\gabriel.rodrigues\Desktop\lista_caminhos_2.xlsx')
# Obtém as duas colunas do DataFrame
coluna1 = df['nome_cliente']
coluna2 = df['cota']
coluna3 = df2['caminho_arquivo']
coluna4 = df2['numero']

# Cria uma nova lista intercalando as informações das colunas 1 e 2
nova_lista = []
nova_lista2 = []

for valor1, valor2 in zip(coluna1, coluna2):
    nova_lista.append(str(valor1))
    valor3 = str(valor2)
    nova_lista.append(str(valor3.replace('.0','')))
    
for valor1, valor2 in zip(coluna3, coluna4):
    nova_lista2.append(str(valor1))
    valor3 = str(valor2)
    nova_lista2.append(str(valor3.replace('.0',''))) 
x = []
y = []
z = []
w = []
w2 = []         

contador = 1   
dataframe = R'C:\Users\gabriel.rodrigues\Desktop\lista_apos_busca_park.xlsx' 
dataframe2 = R'C:\Users\gabriel.rodrigues\Desktop\lista_nao_encontrado_apos_busca_park.xlsx'
arquivo_excel = openpyxl.load_workbook(dataframe)
arquivo_excel2 = openpyxl.load_workbook(dataframe2)
nome_planilha = 'Planilha1'
nome_planilha2 = 'Planilha2'
planilha = arquivo_excel[nome_planilha]
planilha2 = arquivo_excel2[nome_planilha2]
linha = planilha.max_row + 1 
linha2 = planilha2.max_row + 1 
coluna = 1       
coluna2 = 1                      
print('possui {} documentos pdf para analisar'.format(len(nova_lista2)))
for i in range(0, len(nova_lista2), 2):
    elemento = nova_lista2[i]
    numero = nova_lista2[i+1]
    print('verificando o {}º arquivo '.format(contador))
    elemento_buscado = verificar_correspondencia(elemento, numero)
    contador = contador + 1
    if elemento_buscado:
        
        lista_vazia = elemento_buscado.split('+')
        
        #print(elemento_buscado + '\n encontrado no banco de dados \n')
        lista_vazia.append(' ENCONTRADO')
        lista_vazia.append(elemento)
        for celula in lista_vazia:
            planilha.cell(row=linha, column=coluna).value = celula
            coluna += 1
        arquivo_excel.save(R'C:\Users\gabriel.rodrigues\Desktop\lista_apos_busca_park.xlsx')
       
        
        df2 = df2[df2['caminho_arquivo'] != elemento]
       
        df2.to_excel(R'C:\Users\gabriel.rodrigues\Desktop\lista_caminhos_2.xlsx', index=False)
        
        print('encontrado')
    else:
        lista_vazia2=[]
        print(elemento_buscado)
        lista_vazia2.append(elemento) 
        for celula in lista_vazia2:
            planilha2.cell(row=linha2, column=coluna2).value = celula
            coluna2 += 1
        arquivo_excel2.save(R'C:\Users\gabriel.rodrigues\Desktop\lista_nao_encontrado_apos_busca_park.xlsx')
        
        print(elemento)
        df2 = df2[df2['caminho_arquivo'] != elemento]
        df2.to_excel(R'C:\Users\gabriel.rodrigues\Desktop\lista_caminhos_2.xlsx', index=False)
        print('nao encontrado')
    

  